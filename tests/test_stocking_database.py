from pathlib import Path
import json
import sqlite3
import sys
from io import BytesIO

import pytest
from openpyxl import Workbook

sys.path.insert(0, str(Path(__file__).parents[1] / "scripts"))
from stocking_database import StockingEvent, event_key, import_source, export_json
from bootstrap_archive_csv import extract_workbook_rows, validate_records, validate_summary

FIXTURE = Path(__file__).parent / "fixtures" / "archive_sample.html"


def test_key_is_stable_and_normalized():
    a = StockingEvent("Twin  Lakes Reservoir", "2014-06-14", "Rainbow Trout", 10000, 10.2)
    b = StockingEvent("TWIN-LAKES RESERVOIR", "2014-06-14", "rainbow trout", 10000, 10.2)
    assert event_key(a) == event_key(b)


def test_archive_import_deduplicates_and_preserves_sources(tmp_path):
    db = tmp_path / "stocking.sqlite3"
    result = import_source(db, "archive", "fixture", fixture=FIXTURE)
    assert result["events_seen"] == 3
    assert result["new_events"] == 2
    with sqlite3.connect(db) as conn:
        assert conn.execute("select count(*) from stocking_events").fetchone()[0] == 2
        assert conn.execute("select count(*) from event_sources").fetchone()[0] == 3


def test_second_import_is_idempotent(tmp_path):
    db = tmp_path / "stocking.sqlite3"
    import_source(db, "archive", "fixture", fixture=FIXTURE)
    second = import_source(db, "archive", "fixture", fixture=FIXTURE)
    assert second["new_events"] == 0
    assert second["duplicates"] == 3


def test_export_summary(tmp_path):
    db = tmp_path / "stocking.sqlite3"
    out = tmp_path / "events.json"
    import_source(db, "archive", "fixture", fixture=FIXTURE)
    summary = export_json(db, out)
    assert summary["stocking_events"] == 2
    assert summary["events_by_year"] == {"2014": 1, "2015": 1}
    assert json.loads(out.read_text())["summary"]["earliest_date"] == "2014-06-14"


def workbook_blob(rows):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2014"
    for row in rows:
        sheet.append(row)
    blob = BytesIO()
    workbook.save(blob)
    return blob.getvalue()


def test_workbook_uses_neighboring_water_name_when_link_label_is_atlas():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "2014"
    sheet.append(["Date", "Region", "Water", "Map"])
    sheet.append(["06/14/2014", "northeast", "Wrights Lake", "Atlas"])
    sheet["D2"].hyperlink = (
        "https://ndismaps.nrel.colostate.edu/index.html?app=FishingAtlas&value=680"
    )
    blob = BytesIO()
    workbook.save(blob)

    rows, diagnostics = extract_workbook_rows(blob.getvalue())
    assert len(rows) == 1
    event, sheet_title, row_number, raw, atlas_id, atlas_url = rows[0]
    assert sheet_title == "2014"
    assert row_number == 2
    assert event.water_name == "Wrights Lake"
    assert event.stocking_date == "2014-06-14"
    assert event.region == "northeast"
    assert atlas_id == 680
    assert raw["atlas_id"] == 680
    assert "value=680" in atlas_url
    assert diagnostics[0]["rows_imported"] == 1
    assert diagnostics[0]["atlas_rows_missing_water_name"] == 0


def test_workbook_still_accepts_descriptive_hyperlink_label():
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["06/14/2014", "northeast", "Wrights Lake"])
    sheet["C1"].hyperlink = "https://example.test/?value=680"
    blob = BytesIO()
    workbook.save(blob)

    rows, _ = extract_workbook_rows(blob.getvalue())
    assert rows[0][0].water_name == "Wrights Lake"


def test_record_validation_rejects_generic_atlas_water_name():
    record = (
        StockingEvent("Atlas", "2014-06-14", "Rainbow Trout"),
        "2014",
        2,
        {},
        680,
        "https://example.test/?value=680",
    )
    with pytest.raises(RuntimeError, match="generic Atlas hyperlink label"):
        validate_records([record])


def test_archive_validation_rejects_incomplete_year_range():
    with pytest.raises(RuntimeError, match="expected years"):
        validate_summary(
            {
                "events_by_year": {"2025": 514},
                "stocking_events": 514,
                "unique_waters": 1,
            }
        )


def test_archive_validation_rejects_single_water_snapshot():
    with pytest.raises(RuntimeError, match="unique waters"):
        validate_summary(
            {
                "events_by_year": {str(year): 50 for year in range(2014, 2026)},
                "stocking_events": 600,
                "unique_waters": 1,
            }
        )

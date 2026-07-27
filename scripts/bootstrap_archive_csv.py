#!/usr/bin/env python3
"""Build the fixed 2014-2025 CPW stocking archive from the published XLSX workbook.

The workbook is treated as a one-time historical source. Each imported record uses
the official stocking date, water name, and Fishing Atlas link/ID. Coordinates and
other water metadata remain the responsibility of the Atlas pipeline.
"""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook

from stocking_database import (
    ARCHIVE_URL,
    USER_AGENT,
    StockingEvent,
    event_key,
    export_json,
    init_db,
    upsert,
    utc_now,
)

FIRST_ARCHIVE_YEAR = 2014
LAST_ARCHIVE_YEAR = 2025
REGIONS = {"northeast", "northwest", "southeast", "southwest"}
DATE_FORMATS = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y")
HYPERLINK_FORMULA_RE = re.compile(
    r'^=HYPERLINK\(\s*"([^"]+)"\s*[,;]\s*"([^"]*)"\s*\)$', re.I
)
GENERIC_LINK_LABELS = {"atlas", "map", "link", "fishing atlas", "view atlas"}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def workbook_url(published_url: str) -> str:
    base = published_url.split("?", 1)[0]
    if base.endswith("/pubhtml"):
        base = base[: -len("/pubhtml")] + "/pub"
    elif not base.endswith("/pub"):
        base = base.rstrip("/") + "/pub"
    return f"{base}?output=xlsx"


def atlas_id_from_url(url: str) -> int | None:
    try:
        return int(parse_qs(urlparse(url).query).get("value", [None])[0])
    except (TypeError, ValueError):
        return None


def parse_date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def hyperlink_from_cell(cell) -> tuple[str | None, str]:
    """Return a hyperlink target and its displayed label from an XLSX cell."""
    display = clean(cell.value)
    if cell.hyperlink and cell.hyperlink.target:
        return clean(cell.hyperlink.target), display
    if isinstance(cell.value, str):
        match = HYPERLINK_FORMULA_RE.match(cell.value.strip())
        if match:
            return clean(match.group(1)), clean(match.group(2))
    return None, display


def plausible_water_name(value: object) -> bool:
    text = clean(value)
    lowered = text.casefold()
    if not text or lowered in REGIONS or lowered in GENERIC_LINK_LABELS:
        return False
    if parse_date_value(value) is not None:
        return False
    if lowered.startswith(("http://", "https://", "=hyperlink(")):
        return False
    if re.fullmatch(r"[\d.,%+-]+", text):
        return False
    return bool(re.search(r"[A-Za-z]", text))


def water_name_from_row(row, link_index: int, link_label: str) -> str:
    """Choose the water-name cell associated with an Atlas hyperlink.

    Published CPW workbooks commonly display the hyperlink itself as ``Atlas`` and
    place the actual water name in the immediately preceding cell. Prefer nearby
    cells, falling back to the hyperlink label only when it is descriptive.
    """
    nearby_indexes = [link_index - 1, link_index + 1, link_index - 2, link_index + 2]
    for index in nearby_indexes:
        if 0 <= index < len(row) and plausible_water_name(row[index].value):
            return clean(row[index].value)

    for index, cell in enumerate(row):
        if index != link_index and plausible_water_name(cell.value):
            return clean(cell.value)

    return clean(link_label) if plausible_water_name(link_label) else ""


def extract_workbook_rows(blob: bytes) -> tuple[list[tuple[StockingEvent, str, int, dict, int, str]], list[dict]]:
    workbook = load_workbook(BytesIO(blob), read_only=False, data_only=False)
    extracted: list[tuple[StockingEvent, str, int, dict, int, str]] = []
    diagnostics: list[dict] = []

    for sheet in workbook.worksheets:
        sheet_rows = 0
        dates_seen = 0
        atlas_links_seen = 0
        missing_names = 0

        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            stocking_date = next((parsed for cell in row if (parsed := parse_date_value(cell.value))), None)
            if stocking_date is not None:
                dates_seen += 1
            if stocking_date is None or not (FIRST_ARCHIVE_YEAR <= stocking_date.year <= LAST_ARCHIVE_YEAR):
                continue

            atlas_url = None
            atlas_id = None
            water_name = ""
            for link_index, cell in enumerate(row):
                candidate_url, link_label = hyperlink_from_cell(cell)
                candidate_id = atlas_id_from_url(candidate_url or "")
                if candidate_id is not None:
                    atlas_url = candidate_url
                    atlas_id = candidate_id
                    water_name = water_name_from_row(row, link_index, link_label)
                    atlas_links_seen += 1
                    break

            if atlas_id is None:
                continue
            if not water_name:
                missing_names += 1
                continue

            values = [clean(cell.value) for cell in row]
            region = next((value.lower() for value in values if value.lower() in REGIONS), None)
            event = StockingEvent(
                water_name=water_name,
                stocking_date=stocking_date.isoformat(),
                region=region,
            )
            raw = {
                "sheet": sheet.title,
                "row": row_number,
                "cells": values,
                "atlas_id": atlas_id,
                "atlas_url": atlas_url,
            }
            extracted.append((event, sheet.title, row_number, raw, atlas_id, atlas_url or ""))
            sheet_rows += 1

        diagnostics.append(
            {
                "sheet": sheet.title,
                "rows_imported": sheet_rows,
                "date_rows_seen": dates_seen,
                "atlas_links_seen": atlas_links_seen,
                "atlas_rows_missing_water_name": missing_names,
            }
        )

    unique: list[tuple[StockingEvent, str, int, dict, int, str]] = []
    seen: set[tuple[str, str, int]] = set()
    for record in extracted:
        event, _, _, _, atlas_id, _ = record
        key = (event.stocking_date, event.water_name.casefold(), atlas_id)
        if key not in seen:
            seen.add(key)
            unique.append(record)
    return unique, diagnostics


def validate_records(records: list[tuple[StockingEvent, str, int, dict, int, str]]) -> None:
    names = {record[0].water_name.casefold() for record in records}
    atlas_ids = {record[4] for record in records}
    if names & GENERIC_LINK_LABELS:
        raise RuntimeError("Snapshot validation failed: a generic Atlas hyperlink label was imported as a water name")
    if len(names) < 25:
        raise RuntimeError(f"Snapshot validation failed: only {len(names)} unique water names were extracted")
    if len(atlas_ids) < 25:
        raise RuntimeError(f"Snapshot validation failed: only {len(atlas_ids)} unique Atlas IDs were extracted")


def ensure_atlas_columns(conn: sqlite3.Connection) -> None:
    existing = {row[1] for row in conn.execute("PRAGMA table_info(stocking_events)")}
    if "atlas_id" not in existing:
        conn.execute("ALTER TABLE stocking_events ADD COLUMN atlas_id INTEGER")
    if "atlas_url" not in existing:
        conn.execute("ALTER TABLE stocking_events ADD COLUMN atlas_url TEXT")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_stocking_atlas_id ON stocking_events(atlas_id)")


def bootstrap(db: Path, published_url: str) -> dict:
    source_url = workbook_url(published_url)
    session = requests.Session()
    session.headers.update({"User-Agent": USER_AGENT})
    response = session.get(source_url, timeout=180)
    response.raise_for_status()
    if len(response.content) < 1024 or not response.content.startswith(b"PK"):
        raise RuntimeError(
            "Published workbook export did not return a valid XLSX file "
            f"(content-type={response.headers.get('content-type')!r}, bytes={len(response.content)})"
        )

    records, diagnostics = extract_workbook_rows(response.content)
    if not records:
        raise RuntimeError(
            "Workbook contained no rows with a 2014-2025 stocking date and a Fishing Atlas hyperlink. "
            f"Sheet diagnostics: {json.dumps(diagnostics)}"
        )
    validate_records(records)

    years = sorted({int(record[0].stocking_date[:4]) for record in records})
    missing_years = [year for year in range(FIRST_ARCHIVE_YEAR, LAST_ARCHIVE_YEAR + 1) if year not in years]
    if years[0] != FIRST_ARCHIVE_YEAR or missing_years:
        raise RuntimeError(
            f"Historical workbook coverage is incomplete. Years found: {years}; missing: {missing_years}; "
            f"sheet diagnostics: {json.dumps(diagnostics)}"
        )

    db.parent.mkdir(parents=True, exist_ok=True)
    if db.exists():
        db.unlink()
    conn = sqlite3.connect(db)
    init_db(conn)
    ensure_atlas_columns(conn)
    observed = utc_now()
    run_id = conn.execute(
        "INSERT INTO import_runs(started_at,source_kind,source_url) VALUES(?,?,?)",
        (observed, "archive-snapshot", source_url),
    ).lastrowid

    new_events = 0
    try:
        for event, sheet_title, row_number, raw, atlas_id, atlas_url in records:
            if upsert(
                conn,
                event,
                source_kind="archive-snapshot",
                source_url=source_url,
                source_sheet=sheet_title,
                source_gid=sheet_title,
                source_row=row_number,
                raw=raw,
                observed_at=observed,
            ):
                new_events += 1
            conn.execute(
                "UPDATE stocking_events SET atlas_id=?, atlas_url=? WHERE event_id=?",
                (atlas_id, atlas_url, event_key(event)),
            )

        conn.execute(
            """UPDATE import_runs SET finished_at=?,rows_seen=?,canonical_events_seen=?,
               new_events=?,duplicate_events=?,errors_json=? WHERE run_id=?""",
            (utc_now(), len(records), len(records), new_events, len(records) - new_events, "[]", run_id),
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "source": "archive-snapshot",
        "source_url": source_url,
        "years_imported": years,
        "rows_seen": len(records),
        "unique_water_names": len({record[0].water_name.casefold() for record in records}),
        "unique_atlas_ids": len({record[4] for record in records}),
        "new_events": new_events,
        "duplicates": len(records) - new_events,
        "sheet_diagnostics": diagnostics,
    }


def validate_summary(summary: dict) -> None:
    years = sorted(int(year) for year in summary.get("events_by_year", {}))
    expected = list(range(FIRST_ARCHIVE_YEAR, LAST_ARCHIVE_YEAR + 1))
    if years != expected:
        raise RuntimeError(f"Snapshot validation failed: expected years {expected}, found {years}")
    if summary.get("stocking_events", 0) <= 514:
        raise RuntimeError(
            f"Snapshot validation failed: only {summary.get('stocking_events', 0)} events were imported"
        )
    if summary.get("unique_waters", 0) < 25:
        raise RuntimeError(
            f"Snapshot validation failed: only {summary.get('unique_waters', 0)} unique waters were imported"
        )


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, default=Path("data/stocking.sqlite3"))
    parser.add_argument("--json", type=Path, default=Path("data/stocking_events.json"))
    parser.add_argument("--archive-url", default=ARCHIVE_URL)
    args = parser.parse_args()

    result = bootstrap(args.db, args.archive_url)
    summary = export_json(args.db, args.json)
    validate_summary(summary)
    print(json.dumps([result, {"export": summary}], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

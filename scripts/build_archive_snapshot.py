#!/usr/bin/env python3
"""Build the fixed 2014-2025 stocking snapshot from CPW's published workbook."""
from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import date, datetime
from io import BytesIO
from pathlib import Path
from urllib.parse import parse_qs, urlparse

import requests
from openpyxl import load_workbook

from stocking_database import (
    ARCHIVE_URL,
    USER_AGENT,
    StockingEvent,
    event_key,
    export_json,
    init_db,
    upsert,
    utc_now,
)

FIRST_ARCHIVE_YEAR = 2014
LAST_ARCHIVE_YEAR = 2025
REGIONS = {"northeast", "northwest", "southeast", "southwest"}
DATE_FORMATS = ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d", "%m-%d-%Y")
HYPERLINK_FORMULA_RE = re.compile(
    r'^=HYPERLINK\(\s*"([^"]+)"\s*[,;]\s*"([^"]*)"\s*\)$', re.I
)
GENERIC_LABELS = {
    "atlas", "water", "water name", "date", "stocking date", "region",
    "fishing atlas", "link", "map", "report",
}


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def workbook_url(published_url: str) -> str:
    base = published_url.split("?", 1)[0]
    if base.endswith("/pubhtml"):
        base = base[: -len("/pubhtml")] + "/pub"
    elif not base.endswith("/pub"):
        base = base.rstrip("/") + "/pub"
    return f"{base}?output=xlsx"


def atlas_id_from_url(url: str) -> int | None:
    try:
        return int(parse_qs(urlparse(url).query).get("value", [None])[0])
    except (TypeError, ValueError):
        return None


def parse_date_value(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = clean(value)
    if not text:
        return None
    for fmt in DATE_FORMATS:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    return None


def hyperlink_from_cell(cell) -> tuple[str | None, str]:
    display = clean(cell.value)
    if cell.hyperlink and cell.hyperlink.target:
        return clean(cell.hyperlink.target), display
    if isinstance(cell.value, str):
        match = HYPERLINK_FORMULA_RE.match(cell.value.strip())
        if match:
            return clean(match.group(1)), clean(match.group(2))
    return None, display


def plausible_water_name(value: object) -> bool:
    text = clean(value)
    if not text or text.casefold() in GENERIC_LABELS or text.casefold() in REGIONS:
        return False
    if parse_date_value(value) is not None:
        return False
    if atlas_id_from_url(text) is not None:
        return False
    return any(character.isalpha() for character in text)


def water_name_from_row(row, link_index: int, link_display: str) -> str:
    """Use the neighboring workbook cell when the hyperlink is merely labelled Atlas."""
    if plausible_water_name(link_display):
        return clean(link_display)

    # The CPW workbook places the water name immediately before its Atlas link.
    # Search outward from that link so harmless blank/formatting columns are tolerated.
    indexes = list(range(link_index - 1, -1, -1)) + list(range(link_index + 1, len(row)))
    for index in indexes:
        value = row[index].value
        if plausible_water_name(value):
            return clean(value)
    return ""


def extract_workbook_rows(blob: bytes):
    workbook = load_workbook(BytesIO(blob), read_only=False, data_only=False)
    extracted = []
    diagnostics = []

    for sheet in workbook.worksheets:
        sheet_rows = dates_seen = atlas_links_seen = 0
        for row_number, row in enumerate(sheet.iter_rows(), start=1):
            stocking_date = next((d for cell in row if (d := parse_date_value(cell.value))), None)
            if stocking_date is not None:
                dates_seen += 1
            if stocking_date is None or not (FIRST_ARCHIVE_YEAR <= stocking_date.year <= LAST_ARCHIVE_YEAR):
                continue

            atlas_url = None
            atlas_id = None
            water_name = ""
            for index, cell in enumerate(row):
                candidate_url, candidate_display = hyperlink_from_cell(cell)
                candidate_id = atlas_id_from_url(candidate_url or "")
                if candidate_id is None:
                    continue
                atlas_url = candidate_url
                atlas_id = candidate_id
                water_name = water_name_from_row(row, index, candidate_display)
                atlas_links_seen += 1
                break

            if atlas_id is None or not water_name:
                continue

            values = [clean(cell.value) for cell in row]
            region = next((value.lower() for value in values if value.lower() in REGIONS), None)
            event = StockingEvent(
                water_name=water_name,
                stocking_date=stocking_date.isoformat(),
                region=region,
            )
            raw = {
                "sheet": sheet.title,
                "row": row_number,
                "cells": values,
                "atlas_id": atlas_id,
                "atlas_url": atlas_url,
            }
            extracted.append((event, sheet.title, row_number, raw, atlas_id, atlas_url or ""))
            sheet_rows += 1

        diagnostics.append({
            "sheet": sheet.title,
            "rows_imported": sheet_rows,
            "date_rows_seen": dates_seen,
            "atlas_links_seen": atlas_links_seen,
        })

    unique = []
    seen = set()
    for record in extracted:
        event, _, _, _, atlas_id, _ = record
        key = (event.stocking_date, event.water_name.casefold(), atlas_id)
        if key not in seen:
            seen.add(key)
            unique.append(record)
    return unique, diagnostics


def ensure_atlas_columns(conn: sqlite3.Connection) -> None:
    existing = {row[1] for row in conn.execute("PRAGMA table_info(stocking_events)")}
    if "atlas_id" not in existing:
        conn.execute("ALTER TABLE stocking_events ADD COLUMN atlas_id INTEGER")
    if "atlas_url" not in existing:
        conn.execute("ALTER TABLE stocking_events ADD COLUMN atlas_url TEXT")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_stocking_atlas_id ON stocking_events(atlas_id)")


def bootstrap(db: Path, published_url: str) -> dict:
    source_url = workbook_url(published_url)
    response = requests.get(source_url, headers={"User-Agent": USER_AGENT}, timeout=180)
    response.raise_for_status()
    if len(response.content) < 1024 or not response.content.startswith(b"PK"):
        raise RuntimeError("Published workbook export did not return a valid XLSX file")

    records, diagnostics = extract_workbook_rows(response.content)
    if not records:
        raise RuntimeError(f"No usable archive records found. Diagnostics: {json.dumps(diagnostics)}")

    years = sorted({int(record[0].stocking_date[:4]) for record in records})
    expected = list(range(FIRST_ARCHIVE_YEAR, LAST_ARCHIVE_YEAR + 1))
    if years != expected:
        raise RuntimeError(f"Historical coverage is incomplete: expected {expected}, found {years}")

    distinct_names = {record[0].water_name.casefold() for record in records}
    if len(distinct_names) < 25 or distinct_names == {"atlas"}:
        raise RuntimeError(
            f"Water-name extraction failed: only {len(distinct_names)} distinct names were found"
        )

    db.parent.mkdir(parents=True, exist_ok=True)
    if db.exists():
        db.unlink()
    conn = sqlite3.connect(db)
    init_db(conn)
    ensure_atlas_columns(conn)
    observed = utc_now()
    run_id = conn.execute(
        "INSERT INTO import_runs(started_at,source_kind,source_url) VALUES(?,?,?)",
        (observed, "archive-snapshot", source_url),
    ).lastrowid

    new_events = 0
    try:
        for event, sheet_title, row_number, raw, atlas_id, atlas_url in records:
            if upsert(
                conn, event,
                source_kind="archive-snapshot",
                source_url=source_url,
                source_sheet=sheet_title,
                source_gid=sheet_title,
                source_row=row_number,
                raw=raw,
                observed_at=observed,
            ):
                new_events += 1
            conn.execute(
                "UPDATE stocking_events SET atlas_id=?, atlas_url=? WHERE event_id=?",
                (atlas_id, atlas_url, event_key(event)),
            )
        conn.execute(
            """UPDATE import_runs SET finished_at=?,rows_seen=?,canonical_events_seen=?,
               new_events=?,duplicate_events=?,errors_json=? WHERE run_id=?""",
            (utc_now(), len(records), len(records), new_events, len(records) - new_events, "[]", run_id),
        )
        conn.commit()
    finally:
        conn.close()

    return {
        "source": "archive-snapshot",
        "source_url": source_url,
        "years_imported": years,
        "rows_seen": len(records),
        "unique_water_names": len(distinct_names),
        "new_events": new_events,
        "duplicates": len(records) - new_events,
        "sheet_diagnostics": diagnostics,
    }


def validate_summary(summary: dict) -> None:
    expected = list(range(FIRST_ARCHIVE_YEAR, LAST_ARCHIVE_YEAR + 1))
    years = sorted(int(year) for year in summary.get("events_by_year", {}))
    if years != expected:
        raise RuntimeError(f"Snapshot validation failed: expected years {expected}, found {years}")
    if summary.get("stocking_events", 0) <= 514:
        raise RuntimeError(f"Snapshot validation failed: only {summary.get('stocking_events', 0)} events")
    if summary.get("unique_waters", 0) < 25:
        raise RuntimeError(f"Snapshot validation failed: only {summary.get('unique_waters', 0)} water names")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--db", type=Path, default=Path("data/stocking.sqlite3"))
    parser.add_argument("--json", type=Path, default=Path("data/stocking_events.json"))
    parser.add_argument("--archive-url", default=ARCHIVE_URL)
    args = parser.parse_args()

    result = bootstrap(args.db, args.archive_url)
    summary = export_json(args.db, args.json)
    validate_summary(summary)
    print(json.dumps([result, {"export": summary}], indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
